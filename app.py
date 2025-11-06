# -*- coding: utf-8 -*-
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import pandas as pd
import os
import datetime

app = Flask(__name__)
app.secret_key = "supersecretkey"

DB = "exam.db"


def init_db():
    with sqlite3.connect(DB) as conn:
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS admins (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        username TEXT UNIQUE,
                        password TEXT
                    )''')
        c.execute('''CREATE TABLE IF NOT EXISTS students (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT,
                        national_id TEXT,
                        phone TEXT,
                        grade TEXT,
                        subject TEXT,
                        score INTEGER,
                        username TEXT UNIQUE,
                        password TEXT
                    )''')
        c.execute('''CREATE TABLE IF NOT EXISTS questions (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        subject TEXT,
                        type TEXT,
                        question TEXT,
                        option1 TEXT,
                        option2 TEXT,
                        option3 TEXT,
                        option4 TEXT,
                        answer TEXT
                    )''')
        c.execute('''CREATE TABLE IF NOT EXISTS results (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        student_id INTEGER,
                        subject TEXT,
                        question_id INTEGER,
                        student_answer TEXT,
                        correct_answer TEXT,
                        is_correct INTEGER,
                        date_taken TEXT
                    )''')
        # إضافة عمود image_path
        c.execute('''CREATE TABLE IF NOT EXISTS exams (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        title TEXT,
                        subject TEXT,
                        duration_minutes INTEGER,
                        code TEXT UNIQUE,
                        is_active INTEGER DEFAULT 1
                    )''')
        # إضافة عمود image_path للأسئلة
        c.execute('''CREATE TABLE IF NOT EXISTS attempts (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        student_id INTEGER,
                        exam_id INTEGER,
                        started_at TEXT,
                        submitted_at TEXT,
                        score INTEGER,
                        UNIQUE(student_id, exam_id)
                    )''')
        conn.commit()

        # إضافة عمود phone للطلاب إن لم يكن موجوداً
        try:
            c.execute("ALTER TABLE questions ADD COLUMN image_path TEXT")
        except Exception:
            pass

        # إضافة عمود status للطلاب إن لم يكن موجوداً
        try:
            c.execute("ALTER TABLE students ADD COLUMN status TEXT")
        except Exception:
            pass

        # إضافة عمود username و password للطلاب إن لم يكن موجوداً
        try:
            c.execute("ALTER TABLE students ADD COLUMN username TEXT UNIQUE")
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE students ADD COLUMN password TEXT")
        except Exception:
            pass
        
        # إضافة عمود section (مثال: قسم أ أو قسم ب) للطلاب إن لم يكن موجوداً
        try:
            c.execute("ALTER TABLE students ADD COLUMN section TEXT")
        except Exception:
            pass
        
        # إضافة عمود name للمشرفين إن لم يكن موجوداً
        try:
            c.execute("ALTER TABLE admins ADD COLUMN name TEXT")
        except Exception:
            pass
        # إضافة عمود user_type للمشرفين (admin/director/system_admin)
        try:
            c.execute("ALTER TABLE admins ADD COLUMN user_type TEXT DEFAULT 'admin'")
        except Exception:
            pass

        # إنشاء جدول نماذج الأسئلة (نموذج 1 ونموذج 2)
        c.execute('''CREATE TABLE IF NOT EXISTS question_models (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        name TEXT,
                        grade TEXT,
                        created_at TEXT
                    )''')
        conn.commit()

        # إضافة عمود model_id للأسئلة
        try:
            c.execute("ALTER TABLE questions ADD COLUMN model_id INTEGER")
        except Exception:
            pass
        # إضافة عمود grade للأسئلة إن لم يكن موجوداً
        try:
            c.execute("ALTER TABLE questions ADD COLUMN grade TEXT")
        except Exception:
            pass
        
        # إضافة عمود grade و model_id و section للامتحانات
        try:
            c.execute("ALTER TABLE exams ADD COLUMN grade TEXT")
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE exams ADD COLUMN model_id INTEGER")
        except Exception:
            pass
        try:
            c.execute("ALTER TABLE exams ADD COLUMN section TEXT")
        except Exception:
            pass
        
        conn.commit()

        # إضافة نماذج افتراضية إن لم تكن موجودة
        c.execute("SELECT COUNT(*) FROM question_models")
        if c.fetchone()[0] == 0:
            c.execute("INSERT INTO question_models (name, grade, created_at) VALUES (?, ?, ?)",
                     ('نموذج 1', '', datetime.datetime.now().isoformat()))
            c.execute("INSERT INTO question_models (name, grade, created_at) VALUES (?, ?, ?)",
                     ('نموذج 2', '', datetime.datetime.now().isoformat()))
            conn.commit()
        
        # إصلاح البيانات: تحديث المشرفين الافتراضيين ليحصلوا على user_type و name
        try:
            c.execute("UPDATE admins SET user_type = 'admin' WHERE user_type IS NULL")
            conn.commit()
        except Exception:
            pass

    create_default_admins()


def create_default_admins():
    admins = [
        ("admin", "1234", "admin"), 
        ("admin1", "1234", "admin"), 
        ("admin2", "1234", "admin"), 
        ("director", "1234", "director"), 
        ("system_admin", "1234", "system_admin")
    ]
    with sqlite3.connect(DB) as conn:
        c = conn.cursor()
        for user, pw, utype in admins:
            try:
                c.execute("INSERT OR IGNORE INTO admins (username, password, user_type) VALUES (?, ?, ?)",
                          (user, generate_password_hash(pw), utype))
            except Exception:
                continue
        conn.commit()
        
        # إضافة محاولات الامتحان للطلاب
        try:
            c.execute("UPDATE admins SET user_type = 'director' WHERE username = 'director'")
            c.execute("UPDATE admins SET user_type = 'system_admin' WHERE username = 'system_admin'")
            conn.commit()
        except Exception:
            pass


init_db()


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        with sqlite3.connect(DB) as conn:
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT * FROM admins WHERE username=?", (username,))
            admin = c.fetchone()
        if admin and check_password_hash(admin['password'], password):
            # التحقق من وجود المستخدم في جدول المشرفين
            try:
                user_type = admin['user_type']
            except (KeyError, IndexError):
                user_type = 'admin'  # افتراضي admin إن لم يكن موجوداً
            if user_type in ['director', 'system_admin']:
                flash("لا يمكنك تسجيل الدخول كمدير من هنا")
                return redirect(url_for('admin_login'))
            session['admin'] = username
            return redirect(url_for('admin_dashboard'))
        else:
            flash("اسم المستخدم أو كلمة المرور غير صحيحة")
    return render_template('admin_login.html')


# تسجيل دخول المدير
@app.route('/director_login', methods=['GET', 'POST'])
def director_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        with sqlite3.connect(DB) as conn:
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT * FROM admins WHERE username=?", (username,))
            admin = c.fetchone()
        if admin and check_password_hash(admin['password'], password):
            try:
                user_type = admin['user_type']
            except (KeyError, IndexError):
                user_type = 'admin'
            if user_type == 'director':
                session['director'] = username
                return redirect(url_for('director_dashboard'))
            else:
                flash("ليس لديك صلاحية الدخول كمدير")
        else:
            flash("اسم المستخدم أو كلمة المرور غير صحيحة")
    return render_template('director_login.html')


# صفحة تسجيل الدخول
@app.route('/system_admin_login', methods=['GET', 'POST'])
def system_admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        with sqlite3.connect(DB) as conn:
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute("SELECT * FROM admins WHERE username=?", (username,))
            admin = c.fetchone()
        if admin and check_password_hash(admin['password'], password):
            try:
                user_type = admin['user_type']
            except (KeyError, IndexError):
                user_type = 'admin'
            if user_type == 'system_admin':
                session['system_admin'] = username
                return redirect(url_for('system_admin_dashboard'))
            else:
                flash("ليس لديك صلاحية الدخول كمدير نظام")
        else:
            flash("اسم المستخدم أو كلمة المرور غير صحيحة")
    return render_template('system_admin_login.html')


# صفحة تسجيل الدخول: للمشرفين والمديرين
@app.route('/system_admin_dashboard')
def system_admin_dashboard():
    if 'system_admin' not in session:
        return redirect(url_for('system_admin_login'))
    
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # الصفحة الرئيسية
    cur.execute("SELECT COUNT(*) FROM exams")
    total_exams = cur.fetchone()[0]
    
    cur.execute("SELECT COUNT(*) FROM questions")
    total_questions = cur.fetchone()[0]
    
    # لوحة تحكم المشرف
    cur.execute("SELECT COUNT(*) FROM admins WHERE user_type = 'admin' OR user_type IS NULL")
    total_admins = cur.fetchone()[0]
    
    cur.execute("SELECT COUNT(*) FROM question_models")
    total_models = cur.fetchone()[0]
    
    # لوحة تحكم المدير
    cur.execute("SELECT COUNT(*) FROM students")
    total_students = cur.fetchone()[0]
    
    # لوحة تحكم المشرف
    cur.execute("SELECT COUNT(*) FROM admins WHERE user_type = 'director'")
    total_directors = cur.fetchone()[0]
    
    # لوحة تحكم مدير النظام
    cur.execute("SELECT COUNT(*) FROM admins WHERE user_type = 'system_admin'")
    total_system_admins = cur.fetchone()[0]
    
    # عدد الطلاب الذين ينتظرون الاعتماد
    cur.execute("SELECT COUNT(*) FROM students WHERE status = 'pending'")
    pending_students_count = cur.fetchone()[0]
    
    # عرض جميع الطلاب (للمشرف)
    cur.execute("SELECT * FROM admins ORDER BY id")
    all_users_rows = cur.fetchall()
    # تحويل sqlite3.Row إلى dict لتمريره إلى Jinja2
    all_users = [dict(zip([col[0] for col in cur.description], row)) for row in all_users_rows]
    
    # جلب جميع الطلاب (مجمعة حسب national_id لتجنب التكرار)
    # نفضل السجل المعتمد إذا كان موجوداً، وإلا نختار أحدث سجل
    # إذا كان national_id NULL أو فارغ، نعرض جميع السجلات
    # عرض جميع الطلاب - بدون تجميع لضمان ظهور جميع الطلاب الجدد
    cur.execute("SELECT * FROM students ORDER BY id DESC")
    all_students_rows = cur.fetchall()
    # تحويل sqlite3.Row إلى dict لتمريره إلى Jinja2
    all_students = [dict(zip([col[0] for col in cur.description], row)) for row in all_students_rows]
    
    # عرض جميع النتائج والطلاب
    username = session['system_admin']
    cur.execute("SELECT * FROM admins WHERE username=?", (username,))
    current_user_row = cur.fetchone()
    # تحويل sqlite3.Row إلى dict لتمريره إلى Jinja2
    current_user = dict(zip([col[0] for col in cur.description], current_user_row)) if current_user_row else None
    
    conn.close()
    
    # إضافة إشعار إذا كان هناك طلاب ينتظرون الاعتماد
    if pending_students_count > 0:
        flash(f"🔔 يوجد {pending_students_count} طالب/طالبة ينتظرون الاعتماد")
    
    return render_template('system_admin_dashboard.html', 
                         total_exams=total_exams,
                         total_questions=total_questions,
                         total_admins=total_admins,
                         total_models=total_models,
                         total_students=total_students,
                         total_directors=total_directors,
                         total_system_admins=total_system_admins,
                         pending_students_count=pending_students_count,
                         all_users=all_users,
                         all_students=all_students,
                         current_user=current_user)


# صفحة إضافة سؤال جديد
@app.route('/system_admin_add_user', methods=['GET', 'POST'])
def system_admin_add_user():
    if 'system_admin' not in session:
        return redirect(url_for('system_admin_login'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        name = request.form.get('name')
        user_type = request.form.get('user_type')
        
        if not name or not username or not password or not user_type:
            flash("الاسم واسم المستخدم وكلمة المرور والنوع مطلوبون")
            return redirect(url_for('system_admin_dashboard'))
        
        # التحقق من وجود المستخدم
        conn = sqlite3.connect(DB)
        cur = conn.cursor()
        cur.execute("SELECT * FROM admins WHERE username=?", (username,))
        existing_user = cur.fetchone()
        
        if existing_user:
            conn.close()
            flash("اسم المستخدم موجود بالفعل")
            return redirect(url_for('system_admin_dashboard'))
        
        # حفظ المستخدم الجديد
        hashed_password = generate_password_hash(password)
        cur.execute("INSERT INTO admins (username, password, name, user_type) VALUES (?, ?, ?, ?)", (username, hashed_password, name, user_type))
        conn.commit()
        conn.close()
        
        if user_type == "director":
            user_type_name = "مدير"
        elif user_type == "system_admin":
            user_type_name = "مدير نظام"
        else:
            user_type_name = "مشرف"
        flash(f"تم إضافة {user_type_name} بنجاح")
        return redirect(url_for('system_admin_dashboard'))
    
    return redirect(url_for('system_admin_dashboard'))


# صفحة عرض الأسئلة
@app.route('/system_admin_delete_user/<int:user_id>', methods=['POST'])
def system_admin_delete_user(user_id):
    if 'system_admin' not in session:
        return redirect(url_for('system_admin_login'))
    
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    
    # حذف المستخدم
    cur.execute("DELETE FROM admins WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    
    flash("تم حذف المستخدم بنجاح")
    return redirect(url_for('system_admin_dashboard'))


# صفحة إضافة الأسئلة
@app.route('/system_admin_update_user_type/<int:user_id>', methods=['POST'])
def system_admin_update_user_type(user_id):
    if 'system_admin' not in session:
        return redirect(url_for('system_admin_login'))
    
    user_type = request.form.get('user_type')
    
    if not user_type:
        flash("يجب تحديد نوع المستخدم")
        return redirect(url_for('system_admin_dashboard'))
    
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    
    cur.execute("UPDATE admins SET user_type = ? WHERE id = ?", (user_type, user_id))
    conn.commit()
    conn.close()
    
    flash("تم تحديث نوع المستخدم بنجاح")
    return redirect(url_for('system_admin_dashboard'))


# صفحة حذف السؤال
@app.route('/system_admin_edit_user/<int:user_id>', methods=['POST'])
def system_admin_edit_user(user_id):
    if 'system_admin' not in session:
        return redirect(url_for('system_admin_login'))
    
    name = request.form.get('name')
    username = request.form.get('username')
    password = request.form.get('password')
    
    if not name or not username:
        flash("الاسم واسم المستخدم مطلوبان")
        return redirect(url_for('system_admin_dashboard'))
    
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    
    cur.execute("SELECT id FROM admins WHERE username = ? AND id != ?", (username, user_id))
    existing_user = cur.fetchone()
    
    if existing_user:
        conn.close()
        flash("اسم المستخدم موجود بالفعل")
        return redirect(url_for('system_admin_dashboard'))
    
    if password:
        hashed_password = generate_password_hash(password)
        cur.execute("UPDATE admins SET name = ?, username = ?, password = ? WHERE id = ?", (name, username, hashed_password, user_id))
    else:
        cur.execute("UPDATE admins SET name = ?, username = ? WHERE id = ?", (name, username, user_id))
    
    conn.commit()
    conn.close()
    
    flash("تم تحديث بيانات المستخدم بنجاح")
    return redirect(url_for('system_admin_dashboard'))


@app.route('/system_admin_add_student', methods=['GET', 'POST'])
def system_admin_add_student():
    if 'system_admin' not in session:
        return redirect(url_for('system_admin_login'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        if not username or not password:
            flash("اسم المستخدم وكلمة المرور مطلوبان")
            return redirect(url_for('system_admin_dashboard'))
        
        # معالجة رفع الصور    
        conn = sqlite3.connect(DB)
        cur = conn.cursor()
        cur.execute("SELECT * FROM students WHERE username=?", (username,))
        existing_student = cur.fetchone()
        
        if existing_student:
            conn.close()
            flash("اسم المستخدم موجود بالفعل")
            return redirect(url_for('system_admin_dashboard'))
        
        # معالجة رفع الصور
        hashed_password = generate_password_hash(password)
        cur.execute("INSERT INTO students (username, password) VALUES (?, ?)", 
                   (username, hashed_password))
        conn.commit()
        conn.close()
        
        flash(f"تم إضافة الطالب {username} بنجاح")
        return redirect(url_for('system_admin_dashboard'))
    
    return redirect(url_for('system_admin_dashboard'))


#  
@app.route('/system_admin_delete_student/<int:student_id>', methods=['POST'])
def system_admin_delete_student(student_id):
    if 'system_admin' not in session:
        return redirect(url_for('system_admin_login'))
    
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    
    #  
    cur.execute("DELETE FROM students WHERE id=?", (student_id,))
    conn.commit()
    conn.close()
    
    flash("تم حذف الطالب بنجاح")
    return redirect(url_for('system_admin_dashboard'))


#  
@app.route('/system_admin_approve_student/<int:student_id>', methods=['POST'])
def system_admin_approve_student(student_id):
    if 'system_admin' not in session:
        return redirect(url_for('system_admin_login'))
    
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    
    # صفحة الطلاب
    cur.execute("UPDATE students SET status = 'approved' WHERE id = ?", (student_id,))
    conn.commit()
    conn.close()
    
    flash("تم قبول الطالب بنجاح")
    return redirect(url_for('system_admin_dashboard'))


#    إرسال SMS
@app.route('/system_admin_delete_all_data', methods=['POST'])
def system_admin_delete_all_data():
    if 'system_admin' not in session:
        return redirect(url_for('system_admin_login'))
    
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    
    try:
        # حذف جميع النتائج
        cur.execute("DELETE FROM questions")
        
        # حذف جميع الامتحانات
        cur.execute("DELETE FROM exams")
        
        # حذف جميع المحاولات
        cur.execute("DELETE FROM attempts")
        
        # حذف جميع النتائج
        cur.execute("DELETE FROM results")
        
        conn.commit()
        conn.close()
        
        flash("تم حذف جميع البيانات بنجاح")
    except Exception as e:
        conn.close()
        flash("حدث خطأ أثناء حذف البيانات")
    
    return redirect(url_for('system_admin_dashboard'))


# لوحة تحكم المدير: عرض جميع النتائج والطلاب والحسابات
@app.route('/director_dashboard')
def director_dashboard():
    if 'director' not in session:
        return redirect(url_for('director_login'))
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # حساب النتائج لكل طالب لكل مادة: النقاط الصحيحة لكل مادة
    cur.execute(
        """
        WITH subject_scores AS (
            SELECT r.student_id,
                   r.subject,
                   SUM(CASE WHEN r.is_correct = 1 THEN 1 ELSE 0 END) AS subject_score
            FROM results r
            GROUP BY r.student_id, r.subject
        )
        SELECT s.id AS student_id,
               s.name AS student_name,
               s.grade,
               COALESCE((SELECT SUM(sc.subject_score) FROM subject_scores sc WHERE sc.student_id = s.id), 0) AS total_score,
               s.status
        FROM students s
        ORDER BY total_score DESC, s.name
        """
    )
    rows = cur.fetchall()
    
    # جلب معلومات المستخدم الحالي
    username = session['director']
    cur.execute("SELECT * FROM admins WHERE username=?", (username,))
    current_user_row = cur.fetchone()
    current_user = dict(zip([col[0] for col in cur.description], current_user_row)) if current_user_row else None
    
    conn.close()

    return render_template('director_dashboard.html', students=rows, current_user=current_user)


def send_sms(phone, message):
    """
    دالة لإرسال رسالة SMS باستخدام خدمة Twilio
    """
    try:
        #  Twilio   
        try:
            from twilio.rest import Client
        except ImportError:
            print("مكتبة Twilio غير مثبتة. سيتم محاكاة الإرسال...")
            print(f"⚠️ [SIMULATED] SMS إلى {phone}: {message}")
            return True
        
        #  Twilio      
        account_sid = os.getenv('TWILIO_ACCOUNT_SID', '')
        auth_token = os.getenv('TWILIO_AUTH_TOKEN', '')
        from_number = os.getenv('TWILIO_FROM_NUMBER', '')
        
        # معالجة رفع الصور  Twilio ɡ   
        if not all([account_sid, auth_token, from_number]):
            print("بيانات اعتماد Twilio غير متوفرة. سيتم محاكاة الإرسال...")
            print(f"⚠️ [SIMULATED] SMS إلى {phone}: {message}")
            return True
        
        #  SMS   Twilio
        client = Client(account_sid, auth_token)
        client.messages.create(
            body=message,
            from_=from_number,
            to=phone
        )
        
        print(f"✅ تم إرسال SMS بنجاح إلى {phone}")
        return True
        
    except Exception as e:
        print(f"❌ خطأ في إرسال SMS: {str(e)}")
        # في حالة الخطأ، يتم محاكاة الإرسال
        print(f"⚠️ [SIMULATED] SMS إلى {phone}: {message}")
        return False


@app.route('/director/accept/<int:student_id>', methods=['POST'])
def director_accept(student_id):
    if 'director' not in session:
        return redirect(url_for('director_login'))
    with sqlite3.connect(DB) as conn:
        c = conn.cursor()
        # معالجة رفع الصور
        c.execute("SELECT name, phone FROM students WHERE id=?", (student_id,))
        student = c.fetchone()
        
        #  
        c.execute("UPDATE students SET status='accepted' WHERE id=?", (student_id,))
        conn.commit()
        
        # معالجة رفع الصورSMS
        if student and student[1]:  #     
            message = f" {student[0]}!      .    ."
            send_sms(student[1], message)
    
    return redirect(url_for('director_dashboard'))


@app.route('/director/reject/<int:student_id>', methods=['POST'])
def director_reject(student_id):
    if 'director' not in session:
        return redirect(url_for('director_login'))
    with sqlite3.connect(DB) as conn:
        c = conn.cursor()
        # معالجة رفع الصور
        c.execute("SELECT name, phone FROM students WHERE id=?", (student_id,))
        student = c.fetchone()
        
        #  
        c.execute("UPDATE students SET status='rejected' WHERE id=?", (student_id,))
        conn.commit()
        
        # معالجة رفع الصورSMS
        if student and student[1]:  #     
            message = f"  {student[0]}.      .   ."
            send_sms(student[1], message)
    
    return redirect(url_for('director_dashboard'))


#   
@app.route('/director_delete_all_results', methods=['POST'])
def director_delete_all_results():
    if 'director' not in session:
        return redirect(url_for('director_login'))
    
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    
    try:
        # حذف جميع النتائج
        cur.execute("DELETE FROM results")
        
        # حذف جميع المحاولات
        cur.execute("DELETE FROM attempts")
        
        # حذف جميع الطلاب
        cur.execute("DELETE FROM students")
        
        conn.commit()
        conn.close()
        
        flash("تم حذف جميع البيانات بنجاح")
    except Exception as e:
        conn.close()
        flash("حدث خطأ أثناء حذف البيانات")
    
    return redirect(url_for('director_dashboard'))


@app.route('/admin_dashboard')
def admin_dashboard():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    with sqlite3.connect(DB) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT id, title, subject, duration_minutes, code, is_active, model_id, grade FROM exams ORDER BY id DESC")
        exams = cur.fetchall()
        cur.execute("SELECT id, name, grade FROM question_models ORDER BY id")
        models = cur.fetchall()
        
        # جلب معلومات المستخدم الحالي
        username = session['admin']
        cur.execute("SELECT * FROM admins WHERE username=?", (username,))
        current_user_row = cur.fetchone()
        current_user = dict(zip([col[0] for col in cur.description], current_user_row)) if current_user_row else None
    return render_template('admin_dashboard.html', exams=exams, models=models, current_user=current_user)


@app.route('/admin_add_question', methods=['GET', 'POST'])
def admin_add_question():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    if request.method == 'POST':
        subject = request.form.get('subject', '').strip()
        qtype = request.form.get('type', '').strip()
        question = request.form.get('question', '')
        option1 = request.form.get('option1', '')
        option2 = request.form.get('option2', '')
        option3 = request.form.get('option3', '')
        option4 = request.form.get('option4', '')
        answer = request.form.get('answer', '')
        model_id = int(request.form.get('model_id', 0) or 0)
        grade = request.form.get('grade', '').strip()
        image_path = None

        # معالجة نوع "صح أو خطأ" - يتم الحصول على الإجابة من النموذج
        if qtype == 'صح أو خطأ':
            answer = request.form.get('answer_tf', '')
        
        # معالجة نوع السؤال من نوع "صورة"
        if qtype == 'صورة' and 'image' in request.files:
            file = request.files['image']
            if file and file.filename:
                filename = secure_filename(file.filename)
                uploads_dir = os.path.join('static', 'uploads')
                try:
                    os.makedirs(uploads_dir, exist_ok=True)
                except Exception:
                    pass
                save_path = os.path.join(uploads_dir, filename)
                file.save(save_path)
                image_path = f"uploads/{filename}"
        with sqlite3.connect(DB) as conn:
            c = conn.cursor()
            # إضافة image_path و model_id و grade للسؤال
            try:
                c.execute('''INSERT INTO questions 
                            (subject, type, question, option1, option2, option3, option4, answer, image_path, model_id, grade)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                          (subject, qtype, question, option1, option2, option3, option4, answer, image_path, model_id, grade))
                conn.commit()
                flash("تم إضافة السؤال بنجاح!")
            except Exception as e:
                # في حالة وجود خطأ
                try:
                    c.execute('''INSERT INTO questions 
                                (subject, type, question, option1, option2, option3, option4, answer, image_path)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                              (subject, qtype, question, option1, option2, option3, option4, answer, image_path))
                    conn.commit()
                    flash("تم إضافة السؤال بنجاح!")
                except Exception as e2:
                    try:
                        c.execute('''INSERT INTO questions 
                                    (subject, type, question, option1, option2, option3, option4, answer)
                                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                                  (subject, qtype, question, option1, option2, option3, option4, answer))
                        conn.commit()
                        flash("تم إضافة السؤال بنجاح!")
                    except Exception as e3:
                        flash(f"حدث خطأ أثناء إضافة السؤال: {str(e3)}")
        return redirect(url_for('admin_add_question'))
    # جلب جميع النماذج
    with sqlite3.connect(DB) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT id, name, grade FROM question_models ORDER BY id")
        models = cur.fetchall()
    return render_template('add_question.html', models=models)


    # صفحة عرض النتائج - عرض جميع الطلاب
@app.route('/admin/view_results')
def view_results():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))

    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # صفحة عرض النتائج - عرض جميع الطلاب
    cur.execute("""
        SELECT 
            s.id,
            s.name,
            s.grade,
            r.subject,
            COUNT(*) AS total,
            SUM(CASE WHEN r.is_correct = 1 THEN 1 ELSE 0 END) AS correct
        FROM results r
        JOIN students s ON r.student_id = s.id
        GROUP BY s.id, s.name, s.grade, r.subject
        ORDER BY s.name, r.subject
    """)
    
    results_data = []
    for row in cur.fetchall():
        results_data.append({
            'student_id': row['id'],
            'student_name': row['name'] or 'غير محدد',
            'grade': row['grade'] or '-',
            'subject': row['subject'] or '-',
            'total_questions': row['total'] or 0,
            'correct_count': row['correct'] or 0
        })
    
    # حساب إحصائيات النتائج
    cur.execute("""
        SELECT 
            COUNT(DISTINCT s.id) AS total_students,
            COUNT(*) AS total_results,
            SUM(CASE WHEN r.is_correct = 1 THEN 1 ELSE 0 END) AS total_correct
        FROM results r
        JOIN students s ON r.student_id = s.id
    """)
    stats_row = cur.fetchone()
    
    stats = {
        'total_students': stats_row['total_students'] if stats_row else 0,
        'total_results': stats_row['total_results'] if stats_row else 0,
        'total_correct': stats_row['total_correct'] if stats_row else 0
    }
    
    conn.close()
    
    return render_template('view_results.html', results=results_data, stats=stats)




@app.route('/export_results', methods=['GET', 'POST'])
def export_results():
    print("="*60)
    print("تم استدعاء دالة export_results")
    print(f"Session: {session}")
    print(f"Admin in session: {'admin' in session}")
    print("="*60)
    
    if 'admin' not in session:
        flash('يرجى تسجيل الدخول أولاً')
        print("غير مصرح للدخول - غير مسجل")
        return redirect(url_for('admin_login'))
    
    print("جار تصدير النتائج...")
    
    try:
        # محاولة استيراد مكتبة openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            print("✅ تم استيراد openpyxl بنجاح")
        except ImportError as ie:
            error_msg = 'مكتبة openpyxl غير مثبتة. يرجى التثبيت: pip install openpyxl'
            print(f"❌ {error_msg}")
            flash(error_msg)
            return redirect(url_for('view_results'))
        
        # الاتصال بقاعدة البيانات
        conn = sqlite3.connect(DB)
        cur = conn.cursor()
        print("✅ تم الاتصال بقاعدة البيانات بنجاح")
        
        # جلب الطلاب
        cur.execute("""
            SELECT 
                s.name,
                s.grade,
                r.subject,
                SUM(CASE WHEN r.is_correct = 1 THEN 1 ELSE 0 END) AS correct,
                COUNT(*) AS total,
                ROUND((SUM(CASE WHEN r.is_correct = 1 THEN 1 ELSE 0 END) * 100.0 / COUNT(*)), 2) AS percentage
            FROM results r
            JOIN students s ON r.student_id = s.id
            GROUP BY s.id, s.name, s.grade, r.subject
            ORDER BY s.name, r.subject
        """)
        
        rows = cur.fetchall()
        conn.close()
        print(f"✅ تم جلب {len(rows)} صف للتصدير")
        
        if not rows:
            flash('لا يوجد نتائج للتصدير')
            return redirect(url_for('view_results'))
        
        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "نتائج الطلاب"
        
        # إضافة البيانات
        headers = ['اسم الطالب', 'الصف', 'المادة', 'عدد الإجابات الصحيحة', 'عدد الأسئلة', 'النسبة المئوية']
        ws.append(headers)
        
        # إضافة بيانات الطلاب للسطر الأول
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # إضافة البيانات
        for row in rows:
            ws.append([
                row[0] if row[0] else 'غير محدد',
                row[1] if row[1] else '-',
                row[2] if row[2] else '-',
                row[3] if row[3] is not None else 0,
                row[4] if row[4] is not None else 0,
                row[5] if row[5] is not None else 0
            ])
        
        # تعديل عرض الأعمدة للصف الأول
        ws.column_dimensions['A'].width = 25  # اسم الطالب
        ws.column_dimensions['B'].width = 15  # الصف
        ws.column_dimensions['C'].width = 20  # المادة
        ws.column_dimensions['D'].width = 20  # عدد الإجابات الصحيحة
        ws.column_dimensions['E'].width = 15  # عدد الأسئلة
        ws.column_dimensions['F'].width = 15  # النسبة المئوية
        
        # تنسيق الخلايا للصف الأول من البيانات
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # إضافة صف فارغ بين الصف الأول والصف الثاني
        conn = sqlite3.connect(DB)
        cur = conn.cursor()
        
        # في حالة عدم وجود نتائج للطلاب في الاختبار
        cur.execute("""
            SELECT DISTINCT
                s.id,
                s.name,
                s.grade
            FROM students s
            JOIN results r ON s.id = r.student_id
            ORDER BY s.name
        """)
        
        all_students = cur.fetchall()
        
        # قائمة المواد الدراسية
        subjects_list = ['الرياضيات', 'اللغة العربية', 'العلوم', 'اللغة الإنجليزية']
        
        # قائمة المواد الدراسية
        ws2 = wb.create_sheet(title="تفاصيل النتائج")
        
        # إضافة البيانات
        headers2 = ['اسم الطالب', 'الصف', 'المادة', 'اللغة العربية', 'العلوم', 'اللغة الإنجليزية', 'الرياضيات', 'المجموع']
        ws2.append(headers2)
        
        # إضافة البيانات للصف الأول
        for student in all_students:
            student_id, student_name, student_grade = student
            student_name = student_name if student_name else 'غير محدد'
            student_grade = student_grade if student_grade else '-'
            
            # جلب نتائج الطالب لكل مادة
            cur.execute("""
                SELECT 
                    r.subject,
                    SUM(CASE WHEN r.is_correct = 1 THEN 1 ELSE 0 END) AS correct,
                    COUNT(*) AS total,
                    ROUND((SUM(CASE WHEN r.is_correct = 1 THEN 1 ELSE 0 END) * 100.0 / COUNT(*)), 2) AS percentage
                FROM results r
                WHERE r.student_id = ?
                GROUP BY r.subject
            """, (student_id,))
            
            subject_results = {row[0]: {'correct': row[1], 'total': row[2], 'percentage': row[3]} 
                              for row in cur.fetchall()}
            
            # إضافة البيانات للصف
            row_data = [student_name, student_grade]
            total_score = 0
            count_subjects = 0
            
            for subject in subjects_list:
                if subject in subject_results:
                    percentage = subject_results[subject]['percentage']
                    row_data.append(percentage)
                    total_score += percentage
                    count_subjects += 1
                else:
                    row_data.append('-')
            
            # إضافة تنسيق للصف الأول
            row_data.append(total_score if count_subjects > 0 else '-')
            row_data.append(round(total_score / count_subjects, 2) if count_subjects > 0 else '-')
            
            ws2.append(row_data)
        
        # إضافة صف فارغ بين الصف الأول والصف الثاني
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, size=12, color="FFFFFF")
        
        # إضافة البيانات للصفوف الأخرى
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # تعديل عرض الأعمدة
        ws2.column_dimensions['A'].width = 25  # اسم الطالب
        ws2.column_dimensions['B'].width = 15  # الصف
        ws2.column_dimensions['C'].width = 15  # المادة
        ws2.column_dimensions['D'].width = 15  # اللغة العربية
        ws2.column_dimensions['E'].width = 15  # العلوم
        ws2.column_dimensions['F'].width = 15  # اللغة الإنجليزية
        ws2.column_dimensions['G'].width = 15  # الرياضيات
        ws2.column_dimensions['H'].width = 15  # المجموع
        
        # إضافة البيانات للصفوف الأخرى
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=8):
            for cell in row:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        conn.close()
        
        # حفظ الملف
        file_path = os.path.join(os.getcwd(), "نتائج_الطلاب.xlsx")
        print(f"✅ تم حفظ الملف في: {file_path}")
        wb.save(file_path)
        
        if not os.path.exists(file_path):
            print("❌ الملف غير موجود!")
            flash('الملف غير موجود')
            return redirect(url_for('view_results'))
        
        print(f"✅ تم إنشاء الملف بنجاح: {file_path}")
        
        # إرسال الملف
        try:
            response = send_file(
                file_path, 
                as_attachment=True, 
                download_name='نتائج_الطلاب.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            print("✅ تم إرسال الملف بنجاح")
            return response
        except Exception as send_error:
            print(f"❌ خطأ في إرسال الملف: {send_error}")
            # إعادة المحاولة بدون mimetype
            return send_file(file_path, as_attachment=True, download_name='نتائج_الطلاب.xlsx')
        
    except Exception as e:
        error_msg = f'حدث خطأ: {str(e)}'
        print(f"❌ {error_msg}")
        import traceback
        print(f"تفاصيل الخطأ:\n{traceback.format_exc()}")
        flash(error_msg)
        return redirect(url_for('view_results'))


@app.route('/logout')
def logout():
    session.pop('admin', None)
    return redirect(url_for('index'))


@app.route('/create_admins')
def create_admins():
    create_default_admins()
    return "✅ تم إنشاء المشرفين الافتراضيين بنجاح!"


subjects = []


@app.route('/student_select_model', methods=['GET', 'POST'])
def student_select_model():
    # معالجة طلب تسجيل الطالب
    return render_template('student_enter.html')


@app.route('/student_enter', methods=['GET', 'POST'])
def student_enter():
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        id_number = request.form.get('id_number', '').strip()
        phone = request.form.get('phone', '').strip()
        grade = request.form.get('grade', '').strip()
        section = request.form.get('section', '').strip()
        
        if not full_name or not id_number or not phone or not grade or not section:
            flash("يرجى ملء جميع الحقول المطلوبة")
            return render_template('student_enter.html')
        
        # حفظ البيانات في قاعدة البيانات - سيتم تعيين الحالة كـ pending
        with sqlite3.connect(DB) as conn:
            cur = conn.cursor()
            cur.execute("INSERT INTO students (name, national_id, phone, grade, section, status) VALUES (?, ?, ?, ?, ?, 'pending')",
                       (full_name, id_number, phone, grade, section))
            conn.commit()
            student_id = cur.lastrowid
        
        # حفظ بيانات الطالب في الجلسة
        session['student_id'] = student_id
        session['student_name'] = full_name
        session['student_grade'] = grade
        session['student_section'] = section
        
        # رسالة نجاح التسجيل
        flash("✅ تم تسجيل البيانات بنجاح! يرجى الانتظار حتى يتم قبولك من قبل المشرف.")
        return render_template('student_enter.html', student_status='pending')
    
    # معالجة GET - عرض صفحة تسجيل الطالب
    if 'student_id' in session:
        with sqlite3.connect(DB) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            cur.execute("SELECT status FROM students WHERE id = ?", (session['student_id'],))
            student = cur.fetchone()
            
            if student and student['status'] == 'approved':
                # الطالب مقبول - السماح بالدخول للامتحان
                return redirect(url_for('student_start_exam'))
            elif student and student['status'] == 'pending':
                # الطالب في انتظار الموافقة
                return render_template('student_enter.html', student_status='pending')
            else:
                # حالة غير معروفة - مسح الجلسة
                session.clear()
    
    return render_template('student_enter.html', student_status=None)


@app.route('/check_student_status')
def check_student_status():
    """API endpoint للتحقق من حالة الطالب"""
    if 'student_id' not in session:
        return jsonify({'status': 'no_session', 'approved': False})
    
    with sqlite3.connect(DB) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT status FROM students WHERE id = ?", (session['student_id'],))
        student = cur.fetchone()
        
        if student:
            return jsonify({'status': student['status'], 'approved': student['status'] == 'approved'})
        else:
            return jsonify({'status': 'not_found', 'approved': False})


@app.route('/student/start_exam')
def student_start_exam():
    if 'student_id' not in session:
        return redirect(url_for('student_select_model'))
    
    # التحقق من حالة الطالب قبل السماح بالدخول
    with sqlite3.connect(DB) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT status FROM students WHERE id = ?", (session['student_id'],))
        student = cur.fetchone()
        if not student or student['status'] != 'approved':
            flash("لم يتم قبولك بعد. يرجى الانتظار حتى يتم قبولك من قبل المدير.")
        return redirect(url_for('student_enter'))
    
    subjects_list = session.get('subjects')
    model_id = session.get('model_id')
    if not subjects_list:
        # إذا لم يتم اختيار مادة من قبل الطالب
        if not model_id:
            with sqlite3.connect(DB) as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute("SELECT * FROM question_models ORDER BY id")
                models = cur.fetchall()
            
            if models:
                import random
                random_model = random.choice(models)
                model_id = random_model['id']
                session['model_id'] = model_id
                flash(f"✅ تم اختيار النموذج بنجاح: {random_model['name']}")
        
        # حفظ بيانات النموذج في الجلسة
        with sqlite3.connect(DB) as conn:
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            if model_id:
                cur.execute("SELECT DISTINCT subject FROM questions WHERE model_id = ? ORDER BY subject", (model_id,))
            else:
                cur.execute("SELECT DISTINCT subject FROM questions ORDER BY subject")
            subjects_list = [r['subject'] for r in cur.fetchall()]
        session['subjects'] = subjects_list
        session['current_subject_index'] = 0
    
    if not subjects_list:
        flash("⚠️ لا توجد أسئلة متاحة في أي مادة للامتحان")
        return redirect(url_for('student_enter'))
    
    idx = session.get('current_subject_index', 0)
    if idx >= len(subjects_list):
        session.clear()
        return "<h2>✅ تم الانتهاء من جميع المواد. شكراً لك!</h2>"
    subject = subjects_list[idx]
    session['current_subject'] = subject
    return redirect(url_for('student_exam', subject=subject))


@app.route('/student_exam/<subject>', methods=['GET', 'POST'])
def student_exam(subject):
    if 'student_id' not in session:
        return redirect(url_for('student_enter'))
    
    # التحقق من حالة الامتحان - منع إعادة إرسال الإجابات
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    
    # التحقق من المواد التي تم إكمالها من قبل الطالب
    cur.execute("SELECT DISTINCT subject FROM results WHERE student_id = ?", (session['student_id'],))
    completed_subjects = [r['subject'] for r in cur.fetchall()]
    
    # التحقق من فهرس المادة الحالية
    idx = session.get('current_subject_index', 0)
    subjects_list = session.get('subjects') or []
    
    # إذا كانت المادة قد تم إكمالها من قبل وتجاوزنا الفهرس المحدد
    if subject in completed_subjects and idx > 0:
        # التحقق من أن المادة لم يتم إكمالها سابقاً
        check_idx = subjects_list.index(subject) if subject in subjects_list else -1
        if check_idx < idx:
            conn.close()
            return "<h2 style='text-align:center; color:red;'>⚠️ لا يمكن إعادة إرسال إجابات هذه المادة</h2>"
    
    duration_minutes = 0
    model_id = session.get('model_id')
    student_grade = session.get('student_grade')
    # الحصول على مدة الامتحان من النموذج
    if model_id and student_grade:
        cur.execute(
            "SELECT id, type, question, option1, option2, option3, option4, answer, image_path FROM questions WHERE subject = ? AND model_id = ? AND grade = ?",
            (subject, model_id, student_grade))
    elif model_id:
        cur.execute(
            "SELECT id, type, question, option1, option2, option3, option4, answer, image_path FROM questions WHERE subject = ? AND model_id = ?",
            (subject, model_id))
    elif student_grade:
        cur.execute(
            "SELECT id, type, question, option1, option2, option3, option4, answer, image_path FROM questions WHERE subject = ? AND grade = ?",
            (subject, student_grade))
    else:
        cur.execute(
            "SELECT id, type, question, option1, option2, option3, option4, answer, image_path FROM questions WHERE subject = ?",
            (subject,))
    rows = cur.fetchall()
    questions = []
    for r in rows:
        questions.append({
            'id': r['id'],
            'type': r['type'],
            'question': r['question'],
            'options': [opt for opt in (r['option1'], r['option2'], r['option3'], r['option4']) if opt],
            'answer': r['answer'],
            'image_path': r['image_path'] if 'image_path' in r.keys() else None
        })
    if request.method == 'POST':
        student_id = session['student_id']
        now = datetime.datetime.now().isoformat(timespec='seconds')
        for q in questions:
            key = f"q{q['id']}"
            student_ans = request.form.get(key)
            if q['type'] == 'صح أو خطأ' or q['type'] == 'صورة' or q['type'] == 'صح أو خطأ':
                is_correct = 1 if (student_ans and q['answer'] and student_ans.strip() == q['answer'].strip()) else 0
            else:
                is_correct = 0
            cur.execute(
                "INSERT INTO results (student_id, subject, question_id, student_answer, correct_answer, is_correct, date_taken) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (student_id, subject, q['id'], student_ans, q['answer'], is_correct, now))
        conn.commit()
        conn.close()
        session['current_subject_index'] = session.get('current_subject_index', 0) + 1
        return redirect(url_for('student_start_exam'))
    progress = int((idx / len(subjects_list)) * 100) if len(subjects_list) > 0 else 0
    conn.close()
    remaining_seconds = None
    return render_template('exam.html', subject=subject, questions=questions, progress=progress, remaining_seconds=remaining_seconds)


# إنشاء امتحان جديد (POST)
@app.route('/admin/exams', methods=['GET'])
def admin_list_exams():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    with sqlite3.connect(DB) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT id, title, subject, duration_minutes, code, is_active FROM exams ORDER BY id DESC")
        exams = cur.fetchall()
    return render_template('admin_dashboard.html', exams=exams)


@app.route('/admin/exams/new', methods=['POST'])
def admin_create_exam():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    title = request.form.get('title', '').strip()
    subject = request.form.get('subject', '').strip()
    duration = int(request.form.get('duration_minutes', '0') or 0)
    grade = request.form.get('grade', '').strip()
    model_id = request.form.get('model_id', '').strip()
    section = request.form.get('section', '').strip()
    if not title or not subject or not grade or not model_id or not section:
        flash('يرجى ملء جميع الحقول المطلوبة')
        return redirect(url_for('admin_dashboard'))
    
    # إنشاء كود الامتحان
    code = f"EXAM-{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}-{model_id}"
    
    with sqlite3.connect(DB) as conn:
        cur = conn.cursor()
        try:
            cur.execute(
                "INSERT INTO exams (title, subject, duration_minutes, code, grade, model_id, section, is_active) VALUES (?, ?, ?, ?, ?, ?, ?, 1)",
                (title, subject, duration, code, grade, int(model_id), section)
            )
            conn.commit()
            exam_id = cur.lastrowid
            flash('✅ تم إنشاء الامتحان بنجاح')
            return redirect(url_for('admin_add_question_to_exam', exam_id=exam_id))
        except Exception as e:
            flash('❌ حدث خطأ في إنشاء الامتحان')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/exams/<int:exam_id>/toggle', methods=['POST'])
def admin_toggle_exam(exam_id):
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    with sqlite3.connect(DB) as conn:
        cur = conn.cursor()
        cur.execute("UPDATE exams SET is_active = CASE WHEN is_active=1 THEN 0 ELSE 1 END WHERE id = ?", (exam_id,))
        conn.commit()
    return redirect(url_for('admin_dashboard'))


# إضافة سؤال للامتحان
@app.route('/admin/exams/<int:exam_id>/add_question', methods=['GET', 'POST'])
def admin_add_question_to_exam(exam_id):
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    
    # التحقق من وجود الامتحان
    with sqlite3.connect(DB) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM exams WHERE id = ?", (exam_id,))
        exam = cur.fetchone()
        if not exam:
            flash('⚠️ الامتحان غير موجود')
            return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        subject = request.form.get('subject', '').strip()
        qtype = request.form.get('type', '').strip()
        question = request.form.get('question', '')
        option1 = request.form.get('option1', '').strip()
        option2 = request.form.get('option2', '').strip()
        option3 = request.form.get('option3', '').strip()
        option4 = request.form.get('option4', '').strip()
        answer = request.form.get('answer', '').strip()
        image_path = None
        
        # معالجة نوع "صح أو خطأ" - يتم الحصول على الإجابة من النموذج
        if qtype == 'صح أو خطأ':
            answer = request.form.get('answer_tf', '').strip()
        
        # معالجة رفع الصور
        if 'image' in request.files:
            file = request.files['image']
            if file and file.filename:
                try:
                    filename = secure_filename(file.filename)
                    uploads_dir = os.path.join(os.path.dirname(__file__), 'static', 'uploads')
                    os.makedirs(uploads_dir, exist_ok=True)
                    save_path = os.path.join(uploads_dir, filename)
                    file.save(save_path)
                    image_path = f"uploads/{filename}"
                except Exception:
                    pass
        
        # إضافة السؤال
        with sqlite3.connect(DB) as conn:
            c = conn.cursor()
            try:
                c.execute('''INSERT INTO questions 
                            (subject, type, question, option1, option2, option3, option4, answer, image_path, model_id, grade)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                          (subject, qtype, question, option1, option2, option3, option4, answer, image_path, exam['model_id'], exam['grade']))
                conn.commit()
                flash('✅ تم إضافة السؤال بنجاح!')
            except Exception as e:
                flash(f'❌ حدث خطأ في إضافة السؤال: {str(e)}')
        
        return redirect(url_for('admin_add_question_to_exam', exam_id=exam_id))
    
    # جلب الأسئلة المتاحة للامتحان
    with sqlite3.connect(DB) as conn:
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM questions WHERE subject = ? AND model_id = ? AND grade = ? ORDER BY id DESC", 
                   (exam['subject'], exam['model_id'], exam['grade']))
        questions = cur.fetchall()
    
    return render_template('add_question_to_exam.html', exam=exam, questions=questions)


# حذف سؤال
@app.route('/admin/questions/<int:question_id>/delete', methods=['POST'])
def admin_delete_question(question_id):
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    
    conn = sqlite3.connect(DB)
    cur = conn.cursor()
    
    try:
        # جلب معلومات السؤال من قاعدة البيانات
        cur.execute("SELECT model_id, grade FROM questions WHERE id = ?", (question_id,))
        question_info = cur.fetchone()
        
        if question_info:
            # حذف السؤال
            cur.execute("DELETE FROM questions WHERE id = ?", (question_id,))
            conn.commit()
            
            # البحث عن exam_id المرتبط بهذا السؤال
            cur.execute("SELECT id FROM exams WHERE model_id = ? AND grade = ? LIMIT 1", 
                       (question_info[0], question_info[1]))
            exam_row = cur.fetchone()
            conn.close()
            
            if exam_row:
                exam_id = exam_row[0]
                flash('✅ تم حذف السؤال بنجاح')
                return redirect(url_for('admin_add_question_to_exam', exam_id=exam_id))
        
        conn.close()
        flash('⚠️ حدث خطأ في حذف السؤال')
        return redirect(url_for('admin_dashboard'))
    except Exception as e:
        conn.close()
        flash('❌ حدث خطأ في حذف السؤال')
        return redirect(url_for('admin_dashboard'))


# تحديث النموذج - تحديث اسم النموذج للامتحان
@app.route('/admin/models/update', methods=['POST'])
def admin_update_model():
    if 'admin' not in session:
        return redirect(url_for('admin_login'))
    model_id = int(request.form.get('model_id', 0))
    grade = request.form.get('grade', '').strip()
    with sqlite3.connect(DB) as conn:
        cur = conn.cursor()
        cur.execute("UPDATE question_models SET grade = ? WHERE id = ?", (grade, model_id))
        conn.commit()
    flash('✅ تم تحديث النموذج بنجاح')
    return redirect(url_for('admin_dashboard'))


if __name__ == '__main__':
    # محاولة استيراد مكتبة openpyxl في حالة فشل الاستيراد السابق
    try:
        from openpyxl import Workbook
        print("✅ تم استيراد openpyxl بنجاح")
    except ImportError:
        print("⚠️ تحذير: مكتبة openpyxl غير مثبتة - سيتم استخدام pandas")
    
    print("\n" + "="*50)
    print("🚀 الخادم يعمل على: http://127.0.0.1:5000")
    print("="*50 + "\n")

    app.run(debug=True)
